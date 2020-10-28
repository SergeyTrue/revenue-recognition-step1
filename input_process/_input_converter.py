from pathlib import Path
import typing
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from ._input_data import InputData


def _get_actual_period_sales_df(input_data: InputData) -> pd.DataFrame:

    initial_df = pd.read_excel(input_data.actual_period_sales, skiprows=8)
    df = initial_df
    rename_dict = {'Customer': 'customer_number',
                   'Unnamed: 1': 'customer_name',
                   'WBS Element': 'wbs_element',
                   'Sales document': 'sales_document',
                   'Customer PO Number': 'customer_po',
                   'Reason for order': 'order_reason',
                   'Region': 'fns',
                   'Invoice document': 'invoice',
                   'Grosp Code as Posted': 'grosp',
                   'RUB': 'ni',
                   'RUB.1': 'cogs'}
    df = df.rename(columns=rename_dict)
    df.fns = df.fns.astype(str).str.replace('RU/', '')
    df['date'] = datetime.today().strftime('%Y%m%d')
    df['action'] = ''
    df['period'] = 'actual'
    df = df[['date', 'fns', 'customer_number', 'customer_name', 'wbs_element', 'customer_po', 'sales_document',
             'order_reason', 'invoice', 'grosp', 'ni', 'cogs', 'action', 'period', ]]
    return df


def _get_historical_sales_df(history) -> pd.DataFrame:
    history_df = pd.read_excel(history)
    history_df['period'] = ''
    history_df['action'] = ''

    history_df = history_df[['date', 'fns', 'customer_number', 'customer_name', 'wbs_element', 'customer_po',
                             'sales_document', 'order_reason', 'invoice', 'grosp', 'ni', 'cogs', 'action', 'period']]
    return history_df


def _get_month_year(input_data: InputData) -> str:
    wb = load_workbook(input_data.actual_period_sales)
    ws = wb.active
    return ws['B1'].value


def get_data_from_input(input_data: InputData, history: Path):
    actual_sales_df = _get_actual_period_sales_df(input_data)
    historical_sales = _get_historical_sales_df(history)
    period = _get_month_year(input_data)
    return actual_sales_df, historical_sales, period

